import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from collections import defaultdict
import json
import sys
import os


def converte_json(json_data):
    dados_convertidos = []

    for feature in json_data:
        feature_data = {
            "feature_name": feature["feature"],
            "feature_tags": feature["tags"],
            "scenarios": [],
        }

        for cenario in feature["cenarios"]:
            feature_data["scenarios"].append({
                "scenario_name": cenario["cenario"],
                "tags": cenario["tags"],
                "has_examples": cenario["tipo"].lower() == "scenario outline",
                "examples_count": cenario["qtd_execucoes"]
            })

        dados_convertidos.append(feature_data)

    return dados_convertidos

def gerar_planilha_regressao(dados, nome_arquivo):
    print(f"Gerando planilha de regressão para o projeto: {nome_arquivo}")
    wb = Workbook()
    ws_principal = wb.active
    ws_principal.title = "Regressão"

    def aba_perfil(ws, dados):
        preenchimento_cabecalho = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
        preenchimento_titulo = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        alinhamento_centralizado = Alignment(horizontal="center", vertical="center")
        fonte_bold = Font(bold=True)
        ws.merge_cells('A1:G1')
        cell = ws['A1']
        cell.value = "Regressão"
        cell.fill = preenchimento_cabecalho
        cell.alignment = alinhamento_centralizado
        cell.font = fonte_bold

        linha = 5
        total_cenarios = 0
        total_cenarios_automatizados = 0
        total_execucoes = 0
        total_execucoes_automatizadas = 0
        for item in dados:
            ws[f"B{linha}"] = "Tag Feature"
            ws[f"C{linha}"] = "Tag do Cenário"
            ws[f"D{linha}"] = item['feature_name']
            ws[f"E{linha}"] = "AUTOMATIZÁVEL"
            ws[f"F{linha}"] = "AUTOMATIZADO"
            ws[f"G{linha}"] = "EXECUÇÕES"

            for col in "BCDEFG":
                c = ws[f"{col}{linha}"]
                c.fill = preenchimento_titulo
                c.font = fonte_bold
                c.alignment = alinhamento_centralizado

            linha += 1

            inicio_linha_feature = linha

            for scenario in item['scenarios']:
                ws[f"C{linha}"] = ", ".join(scenario['tags'])
                ws[f"D{linha}"] = scenario['scenario_name']
                ws[f"E{linha}"].alignment = alinhamento_centralizado
                ws[f"F{linha}"].alignment = alinhamento_centralizado

                if '@automatizar' in scenario['tags']:
                    ws[f"E{linha}"].value = "SIM"
                    ws[f"E{linha}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f"E{linha}"].font = Font(color="006100", bold=True)
                else:
                    ws[f"E{linha}"].value = "NÃO"
                    ws[f"E{linha}"].fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
                    ws[f"E{linha}"].font = Font(color="FF0000", bold=True)

                if '@automatizado' in scenario['tags']:
                    ws[f"E{linha}"].value = "SIM"
                    ws[f"E{linha}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f"E{linha}"].font = Font(color="006100", bold=True)
                    ws[f"F{linha}"].value = "SIM"
                    ws[f"F{linha}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f"F{linha}"].font = Font(color="006100", bold=True)

                    total_cenarios_automatizados += 1
                    total_execucoes_automatizadas += scenario['examples_count']
                else:
                    ws[f"F{linha}"].value = "NÃO"
                    ws[f"F{linha}"].fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
                    ws[f"F{linha}"].font = Font(color="FF0000", bold=True)

                ws[f"G{linha}"] = scenario['examples_count']
                ws[f"G{linha}"].alignment = alinhamento_centralizado

                total_cenarios += 1
                total_execucoes += scenario['examples_count']

                linha += 1

            fim_linha_feature = linha - 1
            if inicio_linha_feature == fim_linha_feature:
                ws[f"B{inicio_linha_feature}"] = ", ".join(item['feature_tags'])
                cell = ws[f"B{inicio_linha_feature}"]
                cell.alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)
                cell.font = Font(bold=True)
            else:
                intervalo = f"B{inicio_linha_feature}:B{fim_linha_feature}"
                ws.merge_cells(intervalo)
                cell = ws[f"B{inicio_linha_feature}"]
                cell.value = ", ".join(item['feature_tags'])
                cell.alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)
                cell.font = Font(bold=True)

        ws["L3"] = "Resumo Geral"
        ws["L3"].fill = preenchimento_cabecalho
        ws["L3"].alignment = alinhamento_centralizado
        ws["L3"].font = fonte_bold

        resumo = {
            "Total de Cenários": total_cenarios,
            "Total de Cenários Automatizados": total_cenarios_automatizados,
            "Cobertura da Automação": "=(M5/M4)",
            "Total de Execuções": total_execucoes,
            "Total de Execuções Automatizadas": total_execucoes_automatizadas,
            "Cobertura Real da Automação": "=(M8/M7)"
        }
        linha_resumo = 4
        for chave, valor in resumo.items():
            ws[f"L{linha_resumo}"] = chave
            ws[f"M{linha_resumo}"] = valor
            if "Cobertura" in chave:
                ws[f"M{linha_resumo}"].number_format = '0.00%'
            linha_resumo += 1
        for coluna in ws.columns:
            max_length = 0
            coluna_letra = get_column_letter(coluna[0].column)
            for cell in coluna:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[coluna_letra].width = max_length + 2

    aba_perfil(ws_principal, dados)

    tags_p = defaultdict(list)
    for feature in dados:
        for scenario in feature['scenarios']:
            for tag in scenario['tags']:
                if tag.startswith("@p_"):
                    tag_limpa = tag[3:]
                    tags_p[tag_limpa].append({
                        "feature_name": feature["feature_name"],
                        "feature_tags": feature["feature_tags"],
                        "scenario": scenario,
                        "examples_count": scenario["examples_count"]
                    })
    for tag, entradas in tags_p.items():
        if len(entradas) < 2:
            continue
        dados_tag = []
        for entrada in entradas:
            existente = next((f for f in dados_tag if f['feature_name'] == entrada['feature_name']), None)
            if not existente:
                dados_tag.append({
                    "feature_name": entrada['feature_name'],
                    "feature_tags": entrada['feature_tags'],
                    "scenarios": [entrada["scenario"]]
                })  
            else:
                existente["scenarios"].append(entrada["scenario"])
        ws_tag = wb.create_sheet(title=tag[:31])
        aba_perfil(ws_tag, dados_tag)   

    preenchimento_cabecalho = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
    preenchimento_titulo = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    alinhamento_centralizado_quebratxt = Alignment(horizontal="center", vertical="center", wrapText=True)
    alinhamento_centralizado = Alignment(horizontal="center", vertical="center")
    fonte_bold = Font(bold=True)

    ws_principal.merge_cells('A1:G1')
    cell = ws_principal['A1']
    cell.value = "Regressão"
    cell.fill = preenchimento_cabecalho
    cell.alignment = alinhamento_centralizado
    cell.font = fonte_bold

    linha = 5

    total_cenarios = 0
    total_cenarios_automatizados = 0
    total_execucoes = 0
    total_execucoes_automatizadas = 0

    for item in dados:
        ws_principal[f"B{linha}"] = "Tag Feature"
        ws_principal[f"C{linha}"] = "Tag do Cenário"
        ws_principal[f"D{linha}"] = item['feature_name']
        ws_principal[f"E{linha}"] = "AUTOMATIZÁVEL"
        ws_principal[f"F{linha}"] = "AUTOMATIZADO"
        ws_principal[f"G{linha}"] = "EXECUÇÕES"

        for col in "BCDEFG":
            c = ws_principal[f"{col}{linha}"]
            c.fill = preenchimento_titulo
            c.font = fonte_bold
            c.alignment = alinhamento_centralizado

        linha += 1

        inicio_linha_feature = linha
        ws_principal[f"B{linha}"].alignment = alinhamento_centralizado_quebratxt

        for scenario in item['scenarios']:
            ws_principal[f"C{linha}"] = ", ".join(scenario['tags'])
            ws_principal[f"D{linha}"] = scenario['scenario_name']
            ws_principal[f"E{linha}"].alignment = alinhamento_centralizado
            ws_principal[f"F{linha}"].alignment = alinhamento_centralizado

            if '@automatizado' in scenario['tags']:
                ws_principal[f"E{linha}"].value = "SIM"
                ws_principal[f"E{linha}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws_principal[f"E{linha}"].font = Font(color="006100", bold=True)
                ws_principal[f"F{linha}"].value = "SIM"
                ws_principal[f"F{linha}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws_principal[f"F{linha}"].font = Font(color="006100", bold=True)

                total_cenarios_automatizados += 1
                total_execucoes_automatizadas += scenario['examples_count']
            elif '@automatizar' in scenario['tags']:
                ws_principal[f"F{linha}"].value = "NÃO"
                ws_principal[f"F{linha}"].fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
                ws_principal[f"F{linha}"].font = Font(color="FF0000", bold=True)
                ws_principal[f"E{linha}"].value = "SIM"
                ws_principal[f"E{linha}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws_principal[f"E{linha}"].font = Font(color="006100", bold=True)
            elif '@manual' in scenario['tags']:
                ws_principal[f"E{linha}"].value = "NÃO"
                ws_principal[f"E{linha}"].fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
                ws_principal[f"E{linha}"].font = Font(color="FF0000", bold=True)
                ws_principal[f"F{linha}"].value = "NÃO"
                ws_principal[f"F{linha}"].fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
                ws_principal[f"F{linha}"].font = Font(color="FF0000", bold=True)
            else:
                ws_principal[f"E{linha}"].value = "INDEFINIDO"
                ws_principal[f"E{linha}"].fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                ws_principal[f"E{linha}"].font = Font(color="9C6500", bold=True)
                ws_principal[f"F{linha}"].value = "INDEFINIDO"
                ws_principal[f"F{linha}"].fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                ws_principal[f"F{linha}"].font = Font(color="9C6500", bold=True)

            ws_principal[f"G{linha}"] = scenario['examples_count']
            ws_principal[f"G{linha}"].alignment = alinhamento_centralizado

            total_cenarios += 1
            total_execucoes += scenario['examples_count']

            linha += 1

        fim_linha_feature = linha - 1
        if inicio_linha_feature == fim_linha_feature:
            ws_principal[f"B{inicio_linha_feature}"] = ", ".join(item['feature_tags'])
            cell = ws_principal[f"B{inicio_linha_feature}"]
            cell.alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)
            cell.font = Font(bold=True)
        else:
            intervalo = f"B{inicio_linha_feature}:B{fim_linha_feature}"
            ws_principal.merge_cells(intervalo)
            cell = ws_principal[f"B{inicio_linha_feature}"]
            cell.value = ", ".join(item['feature_tags'])
            cell.alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)
            cell.font = Font(bold=True)

    ws_principal["L3"] = "Resumo Geral"
    ws_principal["L3"].fill = preenchimento_cabecalho
    ws_principal["L3"].alignment = alinhamento_centralizado
    ws_principal["L3"].font = fonte_bold

    resumo = {
        "Total de Cenários": total_cenarios,
        "Total de Cenários Automatizados": total_cenarios_automatizados,
        "Cobertura da Automação": "=(M5/M4)",
        "Total de Execuções": total_execucoes,
        "Total de Execuções Automatizadas": total_execucoes_automatizadas,
        "Cobertura Real da Automação": "=(M8/M7)"
    }

    linha_resumo = 4
    for chave, valor in resumo.items():
        ws_principal[f"L{linha_resumo}"] = chave
        ws_principal[f"M{linha_resumo}"] = valor
        if "Cobertura" in chave:
            ws_principal[f"M{linha_resumo}"].number_format = '0.00%'
        linha_resumo += 1

    for coluna in ws_principal.columns:
        max_length = 0
        coluna_letra = get_column_letter(coluna[0].column)
        for cell in coluna:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_principal.column_dimensions[coluna_letra].width = max_length + 2

    output_file = f"Regressao_{nome_arquivo}.xlsx"
    wb.save(output_file)
    print(f"Arquivo '{output_file}' criado com sucesso!")
    return output_file


# Função principal para execução via linha de comando
def main():
    # Verificar se os argumentos foram passados
    if len(sys.argv) == 3:
        # Receber parâmetros via argumentos de linha de comando
        json_file = sys.argv[1]
        nome_projeto = sys.argv[2]
        
        try:
            with open(json_file, "r", encoding="utf-8") as f:
                json_data = json.load(f)
            gerar_planilha_regressao(converte_json(json_data), nome_projeto)
        except Exception as e:
            print(f"ERRO: {str(e)}")
            sys.exit(1)
    else:
        # Tentar obter parâmetros via variáveis de ambiente
        json_file = os.environ.get('CAMINHO_ARQUIVO_JSON')
        nome_projeto = os.environ.get('NOME_PROJETO_ANALISADO')
        
        if json_file and nome_projeto:
            try:
                with open(json_file, "r", encoding="utf-8") as f:
                    json_data = json.load(f)
                gerar_planilha_regressao(converte_json(json_data), nome_projeto)
            except Exception as e:
                print(f"ERRO: {str(e)}")
                sys.exit(1)
        else:
            print("ERRO: Parâmetros insuficientes.")
            print("Uso: python planilha_regressao_json_jenkins.py <caminho_arquivo_json> <nome_projeto>")
            print("Ou defina as variáveis de ambiente CAMINHO_ARQUIVO_JSON e NOME_PROJETO_ANALISADO")
            sys.exit(1)


if __name__ == "__main__":
    main()
